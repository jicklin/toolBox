import os
import logging
from openpyxl import Workbook, load_workbook
from decimal import Decimal

logging.basicConfig(level=logging.INFO,  # 控制台打印的日志级别
                    filename='diff_assign.log',
                    filemode='a',  ##模式，有w和a，w就是写模式，每次都会重新写日志，覆盖之前的日志
                    # a是追加模式，默认如果不写的话，就是追加模式
                    format=
                    '%(asctime)s - %(pathname)s[line:%(lineno)d] - %(levelname)s: %(message)s'
                    # 日志格式
                    )


class DiffAssignTool:

    def __init__(self, targetfile, outpath, groupnum):
        self.parcel_code_map = {}
        self.export_list = []
        self._target_file = targetfile
        if outpath is None or len(outpath) == 0:
            self._out_path = os.path.join(os.path.abspath("."), 'diff')
        else:
            self._out_path = outpath
        self._group_num = groupnum
        if not os.path.exists(self._out_path):
            os.mkdir(self._out_path)
        self._process_with_excel(targetfile)


    def _process_with_excel(self, target_file):
        wb = load_workbook(target_file)
        if wb is None:
            raise BaseException('选择的文件不存在')
        sheet = wb.worksheets[0]
        for row in sheet.iter_rows(min_row=2):
            parcel_code = row[0].value
            total_area = row[1].value
            child_area = row[2].value
            code = parcel_code[0:18]
            # if code == parcel_code:
            #     self.code_total_map[code] = value
            #     continue
            if code not in self.parcel_code_map:
                self.parcel_code_map[code] = [
                    RowDto(parcel_code, Decimal(str(total_area)), Decimal(str(child_area)), None, Decimal(str(child_area)))]
            else:
                self.parcel_code_map[code].append(
                    RowDto(parcel_code, Decimal(str(total_area)), Decimal(str(child_area)), None, Decimal(str(child_area))))

        for key in self.parcel_code_map.keys():
            row_list = self.parcel_code_map[key]
            total = row_list[0].total_area
            child_sum = Decimal("0")

            cmpfun = lambda x: float(x.child_area)
            row_list.sort(key=cmpfun,reverse=True)

            for row in row_list:
                child_sum = child_sum + Decimal(str(row.child_area))
            diff = total - child_sum

            if diff == 0:
                self.export_list.extend(row_list)
                continue

            unit = diff / abs(diff)
            num = abs(diff) / Decimal("0.01")
            while num > 0:
                num = self.assign(row_list, num, unit)
            self.export_list.extend(row_list)

    def assign(self, child_list, num, unit):
        for child_row in child_list:
            diff = 0
            if child_row.diff is None:
                diff = unit * Decimal('0.01')
            else:
                diff = child_row.diff + unit * Decimal('0.01')
            child_row.diff = diff
            child_row.changed_area = child_row.child_area + diff
            num = num - 1
            if num == 0:
                return num
        return num

    def create_file(self):
        data_list = self.export_list
        titles = ['标识码','总面积' ,'原面积', '差值', '校正值']
        wb = Workbook()
        sheet = wb.worksheets[0]
        sheet.title = 'assign'
        # 表头
        for hx in range(1, len(titles) + 1):
            sheet.cell(1, hx, titles[hx - 1])
        for row in range(2, len(data_list) + 2):
            sheet.cell(row, 1, data_list[row - 2].paracel_code)
            sheet.cell(row, 2, data_list[row - 2].total_area)
            sheet.cell(row, 3, data_list[row - 2].child_area)
            sheet.cell(row, 4, data_list[row - 2].diff)
            sheet.cell(row, 5, data_list[row - 2].changed_area)
        wb.save(self._out_path + "/" + data_list[0].paracel_code[0:18] + '等标识码.xlsx')


"""
记录下每列的数据
"""


class RowDto:
    def __init__(self, paracel_code, total_area, child_area, diff=None, changed_area=None):
        self.paracel_code = paracel_code
        self.total_area = total_area
        self.child_area = child_area
        self.diff = diff
        self.changed_area = changed_area

    def __repr__(self):
        return repr((self.paracel_code, self.child_area, self.diff, self.changed_value))


if __name__ == '__main__':
    a = Decimal(abs(Decimal('0.07'))) / Decimal("0.01")
    print(type(a))
    print(a > 0)
    a = a - 1
    print(a)
    # print(1.0*Decimal('7'))
    print(1 + Decimal('7'))
    tool = DiffAssignTool('E:\标识码.xlsx',None,18)
    tool.create_file()


